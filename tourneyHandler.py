#!/usr/bin/env python3

from match import Match
from player import Player
from ratings import Ratings
from pool import Pool
from functions import *
import argparse
import pandas as pd
import os
from openpyxl import load_workbook

# Function to initialize a new tournament spreadsheet
def makeTournament():
    name = "Tournament"
    file_name = f"{name}.xlsx"
    full_path = os.path.abspath(file_name)

    if os.path.exists(file_name):
        print(f"The tournament already exists.")
        return

    df = pd.DataFrame(columns=["PlayerName", "PlayerInfo", "Rating", "Include", "GamesPlayed"])

    df.to_excel(file_name, index=False)
    print(f"Tournament has been initialized at {full_path}.")

def addRound():

    try:
        wb = load_workbook("Tournament.xlsx")
    except FileNotFoundError:
        print("Tournament.xlsx not found. Please make sure the file exists.")
        return

    
    number = len(wb.sheetnames)
    sheet_name = f"Round{number}"

    ws = wb.create_sheet(title=sheet_name)

    players = readRoster()
    sortPlayers(players)
    pools = getPools(players)
    for pool in pools:
        pool.insert_2x2_list(ws)
    
    wb.save("Tournament.xlsx")
    print(f"Sheet '{sheet_name}' added and saved to Tournament.xlsx")

def updateRatings():
    ratingHandler = Ratings(readRoster(), aggregateMatches())
    ratingHandler.updateRatings()

    try:
        wb = load_workbook("Tournament.xlsx")
    except FileNotFoundError:
        print("Tournament.xlsx not found. Please make sure the file exists.")
        return

    ws = wb["Sheet1"]
    row = 2
    for player in ratingHandler.players:
        ws[f"C{row}"] = player.rating
        row += 1
    wb.save("Tournament.xlsx")
    

def main():
    parser = argparse.ArgumentParser(description="Tournament Manager Command Line Tool")
    
    parser.add_argument("command", choices=["makeTournament", "addRound", "updateRatings"], help="Command to run")

    # Parse the arguments
    args = parser.parse_args()
    
    if args.command == "makeTournament":
        makeTournament()
    elif args.command == "addRound":
        addRound()
    elif args.command == "updateRatings":
        updateRatings()

if __name__ == "__main__":
    main()
