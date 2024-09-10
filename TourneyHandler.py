#!/usr/bin/env python3

from match import Match
from player import Player
from ratings import Ratings
import argparse
import pandas as pd
import os
from openpyxl import load_workbook


class Pool:
    def __init__(self, player1, player2, player3, player4, id):
        self.player1 = player1
        self.player2 = player2
        self.player3 = player3
        self.player4 = player4
        self.id = id
        self.matches = []
    
    def createMatches(self):
        self.matches.append(Match(self.player1, self.player2, self.player3, self.player4, 1))
        self.matches.append(Match(self.player1, self.player3, self.player2, self.player4, 1))
        self.matches.append(Match(self.player1, self.player4, self.player2, self.player3, 1))
    
    def makeGrid(self):
        col1 = [f"Pool {self.id}", self.player1.id, self.player2.id, self.player3.id, self.player4.id]
        col2 = ["names", self.player1.name, self.player2.name, self.player3.name, self.player4.name]
        grid = [col1, col2]
        for index, match in enumerate(self.matches):
            col = [f"Match {index+1}"]
            names = ", ".join([player.name for player in match.team1])
            col.append(names)
            col.append("vs")
            names = ", ".join([player.name for player in match.team2])
            col.append(names)
            grid.append(col)
            grid.append([f"result{index+1}"])
        return grid
    
def insert_2x2_list(ws, data, start_row):
    start_col = "A"
    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]

    for i in range(len(data)):
        for j in range(len(data[i])):
            col_letter = cols[i]
            ws[f"{col_letter}{start_row + j}"] = data[i][j]

# Function to initialize a new tournament spreadsheet
def makeTournament():
    name = "Tournament"
    file_name = f"{name}.xlsx"
    full_path = os.path.abspath(file_name)

    if os.path.exists(file_name):
        print(f"The tournament already exists.")
        return

    df = pd.DataFrame(columns=["PlayerName", "PlayerInfo", "Rating"])

    df.to_excel(file_name, index=False)
    print(f"Tournament has been initialized at {full_path}.")

def readRoster():
    file_name = "Tournament.xlsx"
    
    if not os.path.exists(file_name):
        print("The tournament file does not exist.")
        return
    
    df = pd.read_excel(file_name)
    players = []
    for index, row in df.iterrows():
        name = row["PlayerName"]
        rating = row["Rating"]
        player = Player(name, index, rating)
        players.append(player)
    return players


def sortPlayers(players):
    # Sort the players based on their ratings
    players.sort(key=lambda x: x.rating, reverse=True)

def getPools(players):
    pools = []
    if len(players) % 4 != 0:
        print("The number of players is not a multiple of 4. Please add or remove players.")
        return pools
    for index, player in enumerate(players):
        if index%4 == 0:
            pool = Pool(players[index], players[index+1], players[index+2], players[index+3], index//4)
            pool.createMatches()
            pools.append(pool)
    return pools

def addRound(n):
    try:
        wb = load_workbook("Tournament.xlsx")
    except FileNotFoundError:
        print("Tournament.xlsx not found. Please make sure the file exists.")
        return

    sheet_name = f"Round{n}"
    
    if sheet_name in wb.sheetnames:
        print(f"Sheet '{sheet_name}' already exists.")
        return

    ws = wb.create_sheet(title=sheet_name)

    players = readRoster()
    sortPlayers(players)
    pools = getPools(players)
    for index, pool in enumerate(pools):
        grid = pool.makeGrid()
        insert_2x2_list(ws, grid, 6*index+1)
    
    wb.save("Tournament.xlsx")
    print(f"Sheet '{sheet_name}' added and saved to Tournament.xlsx")
    

def main():
    parser = argparse.ArgumentParser(description="Tournament Manager Command Line Tool")
    
    parser.add_argument("command", choices=["makeTournament", "addRound"], help="Command to run")

    # Parse the arguments
    args = parser.parse_args()
    
    if args.command == "makeTournament":
        makeTournament()
    elif args.command == "addRound":
        addRound(1)

if __name__ == "__main__":
    main()
