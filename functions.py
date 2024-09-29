from match import Match
from player import Player
from ratings import Ratings
from pool import Pool
import argparse
import pandas as pd
import os
from openpyxl import load_workbook

def readRoster():
    file_name = "Tournament.xlsx"
    
    if not os.path.exists(file_name):
        print("The tournament file does not exist.")
        return
    
    df = pd.read_excel(file_name)
    players = []
    for index, row in df.iterrows():
        name = row["PlayerName"]
        rating = row["Rating"]
        present = False
        if row["Include"] == 1:
            present = True
        player = Player(name, index, rating, present=present)
        players.append(player)
    return players

def createPlayerDict(players):
    player_dict = {}
    for player in players:
        player_dict[player.id] = player
    return player_dict

def getPoolsFromSheet(sheet, players):
    players = createPlayerDict(players)
    num_rows = sheet.max_row
    counter = 1
    pools = []
    while counter < num_rows:
        pool_id = int(sheet[f"A{counter}"].value.split(" ")[1])
        player1 = players[sheet[f"A{counter+1}"].value]
        player2 = players[sheet[f"A{counter+2}"].value]
        player3 = players[sheet[f"A{counter+3}"].value]
        player4 = players[sheet[f"A{counter+4}"].value]
        results = [sheet[f"D{counter+2}"].value, sheet[f"F{counter+2}"].value, sheet[f"H{counter+2}"].value]
        pool = Pool(player1, player2, player3, player4, pool_id)
        pool.createMatches(results=results)
        counter += 6
        pools.append(pool)
    return pools

def aggregateMatches(players):
    try:
        wb = load_workbook("Tournament.xlsx")
    except FileNotFoundError:
        print("Tournament.xlsx not found. Please make sure the file exists.")
        return
    matches = []
    numRounds = len(wb.sheetnames)
    for i in range(1,numRounds):
        sheet_name = f"Round{i}"
        sheet = wb[sheet_name]
        pools = getPoolsFromSheet(sheet, players)
        for pool in pools:
            for match in pool.matches:
                matches.append(match)
    return matches

def sortPlayers(players):
    # Sort the players based on their ratings
    players.sort(key=lambda x: x.rating, reverse=True)

def getPools(allPlayers):
    players = [player for player in allPlayers if player.present==1]
    pools = []
    if len(players) % 4 != 0:
        print("The number of players is not a multiple of 4. Please add or remove players.")
        return pools
    for index, player in enumerate(players):
        if index%4 == 0:
            pool = Pool(players[index], players[index+1], players[index+2], players[index+3], index//4)
            pool.createMatches()
            pools.append(pool)
    return pools